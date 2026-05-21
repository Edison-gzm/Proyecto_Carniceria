from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from sqlalchemy.orm import Session
from database.models.sale import Sale, SaleStatus
from config import REPORTS_DIR, APP_NAME


class ReportService:
    def __init__(self, session: Session):
        self.session = session

    def export_excel(self, date_from: datetime, date_to: datetime) -> Path:
        sales = (
            self.session.query(Sale)
            .filter(Sale.created_at >= date_from, Sale.created_at <= date_to)
            .filter(Sale.status != SaleStatus.CANCELLED)
            .order_by(Sale.created_at)
            .all()
        )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Ventas"

        # Título
        ws.merge_cells("A1:G1")
        ws["A1"] = f"{APP_NAME} — Reporte de Ventas"
        ws["A1"].font = Font(bold=True, size=13)
        ws["A1"].alignment = Alignment(horizontal="center")

        ws["A2"] = f"Período: {date_from.strftime('%d/%m/%Y')} al {date_to.strftime('%d/%m/%Y')}"
        ws["A2"].font = Font(italic=True)

        # Encabezados
        headers = ["Factura", "Fecha", "Cliente", "Cédula/NIT", "Cajero", "Subtotal", "Total"]
        header_fill = PatternFill("solid", fgColor="1F4E79")
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Datos
        for row, sale in enumerate(sales, 5):
            ws.cell(row=row, column=1, value=sale.invoice_number)
            ws.cell(row=row, column=2, value=sale.created_at.strftime("%d/%m/%Y %H:%M"))
            ws.cell(row=row, column=3, value=sale.customer.full_name if sale.customer else "")
            ws.cell(row=row, column=4, value=sale.customer.id_number if sale.customer else "")
            ws.cell(row=row, column=5, value=sale.user.username if sale.user else "")
            ws.cell(row=row, column=6, value=float(sale.subtotal))
            ws.cell(row=row, column=7, value=float(sale.total))

        # Total general
        total_row = len(sales) + 5
        ws.cell(row=total_row, column=6, value="TOTAL").font = Font(bold=True)
        ws.cell(row=total_row, column=7, value=sum(float(s.total) for s in sales)).font = Font(bold=True)

        # Ancho de columnas
        for col in ws.columns:
            max_len = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[col[0].column_letter].width = max(max_len + 2, 12)

        output_path = Path(REPORTS_DIR) / f"ventas_{date_from.strftime('%Y%m%d')}_{date_to.strftime('%Y%m%d')}.xlsx"
        wb.save(str(output_path))
        return output_path